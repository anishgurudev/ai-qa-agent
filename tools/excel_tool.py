# tools/excel_tool.py
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from crewai.tools import tool
import json, os, re

OUTPUT_DIR = "output"
EXCEL_PATH = "output/test_cases.xlsx"

REQUIRED_KEYS = ["test_id", "module", "test_scenario",
                 "test_steps", "expected_result", "priority", "test_type"]

@tool("Write Test Cases to Excel")
def write_excel_tool(test_cases_json: str) -> str:
    """
    Accepts a JSON string of test cases and writes them into a
    formatted Excel file. Handles both array-of-objects and
    array-of-arrays (CSV-like) formats from the LLM.
    """
    # ── Step 1: Clean input ────────────────────────────────────
    raw = test_cases_json.strip()
    raw = re.sub(r"^```(?:json)?\s*", "", raw)
    raw = re.sub(r"\s*```$", "", raw)
    raw = raw.strip()

    # ── Step 2: Try to extract and parse JSON ──────────────────
    # Find outermost [ ... ] array
    match = re.search(r"\[.*\]", raw, re.DOTALL)
    if match:
        raw = match.group(0)

    try:
        data = json.loads(raw)
    except json.JSONDecodeError:
        # LLM may output comma-separated arrays without outer brackets
        # Wrap lines that look like arrays into one big array
        lines = re.findall(r'\[.*?\]', raw, re.DOTALL)
        if lines:
            try:
                data = json.loads(f"[{','.join(lines)}]")
            except json.JSONDecodeError as e:
                return f"ERROR: Invalid JSON. Details: {e}\nInput (first 300 chars): {test_cases_json[:300]}"
        else:
            return f"ERROR: No JSON array found in output.\nInput (first 300 chars): {test_cases_json[:300]}"

    if not isinstance(data, list) or len(data) == 0:
        return "ERROR: JSON must be a non-empty list."

    # ── Step 3: Convert array-of-arrays → array-of-objects ────
    # If first element is a list (not dict), treat row[0] as headers
    if isinstance(data[0], list):
        headers_row = [str(h).strip().lower().replace(" ", "_") for h in data[0]]
        converted = []
        for row in data[1:]:
            if isinstance(row, list):
                obj = {headers_row[i]: row[i] for i in range(min(len(headers_row), len(row)))}
                converted.append(obj)
        data = converted

    if len(data) == 0:
        return "ERROR: No test case rows found after parsing."

    # ── Step 4: Create Workbook ────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Test Cases"

    headers = ["Test ID", "Module", "Test Scenario",
               "Test Steps", "Expected Result", "Priority", "Test Type"]

    BLUE   = PatternFill("solid", fgColor="1F4E79")
    WHITE  = Font(bold=True, color="FFFFFF", size=11)
    CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill  = BLUE
        cell.font  = WHITE
        cell.alignment = CENTER
    ws.row_dimensions[1].height = 30

    PRIORITY_COLORS = {"high": "FF6B6B", "medium": "FFD966", "low": "92D050"}

    for row_num, tc in enumerate(data, 2):
        ws.cell(row=row_num, column=1, value=tc.get("test_id",        f"TC_{row_num-1:03d}"))
        ws.cell(row=row_num, column=2, value=tc.get("module",         "General"))
        ws.cell(row=row_num, column=3, value=tc.get("test_scenario",  ""))
        ws.cell(row=row_num, column=4, value=tc.get("test_steps",     ""))
        ws.cell(row=row_num, column=5, value=tc.get("expected_result",""))
        ws.cell(row=row_num, column=6, value=tc.get("priority",       "Medium"))
        ws.cell(row=row_num, column=7, value=tc.get("test_type",      "Functional"))

        ws.cell(row=row_num, column=4).alignment = Alignment(wrap_text=True)
        ws.cell(row=row_num, column=5).alignment = Alignment(wrap_text=True)

        priority_key = tc.get("priority", "medium").lower()
        color = PRIORITY_COLORS.get(priority_key, "FFFFFF")
        ws.cell(row=row_num, column=6).fill      = PatternFill("solid", fgColor=color)
        ws.cell(row=row_num, column=6).alignment = Alignment(horizontal="center")

    for col_letter, max_w in {"A":10,"B":18,"C":35,"D":45,"E":45,"F":12,"G":18}.items():
        ws.column_dimensions[col_letter].width = max_w
    ws.freeze_panes = "A2"

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    wb.save(EXCEL_PATH)
    return f"SUCCESS: Excel saved at '{EXCEL_PATH}' with {len(data)} test cases written."
# ── Add at the very BOTTOM of excel_tool.py ────────────────
def write_excel(test_cases_json: str) -> str:
    """Plain callable — cleans LLM output before writing Excel."""
    import re

    raw = test_cases_json.strip()

    # Remove markdown fences
    raw = re.sub(r"^```(?:json)?\s*", "", raw)
    raw = re.sub(r"\s*```$", "", raw)
    raw = raw.strip()

    # ── Fix 1: Convert test_steps arrays → newline string ──────
    # Matches "test_steps": ["step1", "step2"] and joins them
    def join_steps_array(match):
        inner = match.group(1)
        # Extract each quoted string inside the array
        items = re.findall(r'"((?:[^"\\]|\\.)*)"', inner)
        joined = "\\n".join(items)
        return f'"test_steps": "{joined}"'

    raw = re.sub(
        r'"test_steps"\s*:\s*\[(.*?)\]',
        join_steps_array,
        raw,
        flags=re.DOTALL
    )

    # ── Fix 2: Remove trailing commas before } or ] ────────────
    raw = re.sub(r',\s*([}\]])', r'\1', raw)

    # ── Fix 3: Extract outermost [ ... ] array ─────────────────
    match = re.search(r'\[.*\]', raw, re.DOTALL)
    if match:
        raw = match.group(0)

    return write_excel_tool.run(raw)
